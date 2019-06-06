# coding=utf-8
from datetime import timedelta, datetime
import pandas as pd
from openpyxl import load_workbook,Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side,PatternFill,Side, Alignment, Protection
import os


def get_file_info():
    path = os.getcwd()
    files_list = []
    for file in os.listdir(path):
        if os.path.splitext(file)[1] == '.xlsx':
            files_list.append(os.path.join(path, file))
    return files_list, path

def get_values():
    files_list, path = get_file_info()
    wb = load_workbook(filename=files_list[-1])
    ws = wb.active

    val = []
    val_dict = {}
    for i in ws.values:
        val.append(i)
    val_dict['date'] = val[0][7:]
    for j in range(2,len(val),2):
        if val[j][5] == None:
            continue
        val_dict[val[j][5]] = val[j][7:]
    val_dict = dict(sorted(val_dict.items(), key=lambda d: d[0]))
    df_val = pd.DataFrame(val_dict).set_index('date')
    df_val = df_val.drop(['TTL'])
    col_name = list(df_val.columns)
    col_name.append('date')

    new_data = []
    for i in range(len(df_val)):
        if sum(df_val.iloc[i]) > 0:
            temp = list(df_val.iloc[i])
            if df_val.index[i].strftime('%w') == '3':
                temp += [df_val.index[i]+timedelta(days=-2)]
            elif df_val.index[i].strftime('%w') in ['1','4','5']:
                temp += [df_val.index[i] + timedelta(days=-3)]
            elif df_val.index[i].strftime('%w') == '2':
                temp += [df_val.index[i] + timedelta(days=-4)]
            new_data.append(temp)
    df = pd.DataFrame(new_data, columns=col_name)
    df = df.groupby('date').sum()
    return df

df = get_values()

date_list = [df.index[0]]
while date_list[-1] < df.index[len(df)-1]:
    date_list.append(date_list[-1] + timedelta(days=1))

title_row = ['P#','']
for i in date_list:
    title_row.append(i.strftime('%m-%d'))

col_list = list(df.columns)
rows = []

for j in col_list:
    temp_row = [j,'深超需求']
    temp_row2 = ['', '無錫回覆']
    temp_row3 = ['', '备注']
    for k in date_list:
        if k in df.index and df[j][k] != 0:
            temp_row.append(df[j][k])
        else:
            temp_row.append('')
        temp_row2.append('')
        temp_row3.append('')
    rows += [temp_row,temp_row2,temp_row3]
# main()


ft = Font(name=u'宋体',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
fill = PatternFill(fill_type="solid",start_color='FFEEFFFF', end_color='FF001100')
#边框可以选择的值为：'hair', 'medium', 'dashDot', 'dotted', 'mediumDashDot', 'dashed', 'mediumDashed', 'mediumDashDotDot', 'dashDotDot', 'slantDashDot', 'double', 'thick', 'thin']
#diagonal 表示对角线
bd = Border(left=Side(border_style="thin",color='FF001000'),
    right=Side(border_style="thin",color='FF110000'),
    top=Side(border_style="thin",color='FF110000'),
    bottom=Side(border_style="thin",color='FF110000'),
    diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
    vertical=Side(border_style=None,color='FF000000'),
    horizontal=Side(border_style=None,       color='FF110000'))
alignment=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=True,indent=0)
number_format = 'General'
protection = Protection(locked=True,hidden=False)

wx = Workbook()
wc = wx.active
wc.font = ft
wc.fill =fill
wc.border = bd
wc.alignment = alignment
wc.number_format = number_format
wc.title = datetime.today().strftime('%Y%m')

wc.append(title_row)
for row in rows:
    wc.append(row)

for mc in range(2,len(rows) + 2,3):
    col1 = 'A' + str(mc) + ':A' + str(mc+2)
    wc.merge_cells(col1)

row = wc.row_dimensions[1]
row.font = Font(bold=True)   #将A列设定为粗体


wc['c13'] = 'testtest'



# tab = Table(displayName="test1", ref="A1:E5")
#
# # Add a default style with striped rows and banded columns
# style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,\
#                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
# tab.tableStyleInfo = style
# wc.add_table(tab)
wx.save("table3.xlsx")

# wb = load_workbook(filename = 'empty_book.xlsx')
# sheet_ranges = wb['range names']
# print(sheet_ranges['D18'].value)
